import datetime
from openpyxl import load_workbook
import pendulum

def whatistimern():
    now = datetime.datetime.now()
    string = list()
    string.append(str(now.time()).split(":"))
    timeinmin=int(string[0][0])*60+int(string[0][1])
    return timeinmin

def get_date(date):

    if type(date) is list:
        return pendulum.datetime(date[2], date[1], date[0])
    return date

def get_week_parity(date):

    odd_day = pendulum.datetime(2022, 11, 21)
    delta = (get_date(date) - odd_day).in_days()

    return int(delta / 7) % 2


mon = 'Понедельник'
tue = 'Вторник' 
wed = 'Среда'
thur = 'Четверг'
fri = 'Пятница'
sat = 'Суббота'
sun = 'Воскресенье'


days = {
    
    0: mon,
    1: tue,
    2: wed,
    3: thur,
    4: fri,
    5: sat,
    6: sun
}

times = {

    1: '08:00-09:30',
    2: '09:50-11:20',
    3: '11:40-13:10',
    4: '13:40-15:10',
    5: '15:30-17:00',
    6: '17:20-18:50',
    7: '19:05-20:35',
    8: '20:50-22:20'
}


def day_to_num(day):

    if day == mon:
        return 0
    elif day == tue:
        return 1
    elif day == wed:
        return 2
    elif day == thur:
        return 3
    elif day == fri:
        return 4
    elif day == sat:
        return 5

book = load_workbook(filename='website/static/sc/books/fkti-2.xlsx')
sheet = book.active

time, week, day, subject, kafedra, teacher, lesson_type, form, aud, group, course = [i for i in range(11)]

row_start = 4
groups = {}

comps = {}


for row in range(row_start, sheet.max_row + 1):

    gr = sheet[row][group].value
    wk = int(sheet[row][week].value)
    d = str(sheet[row][day].value)
    nd = day_to_num(d)
    lesson_time = sheet[row][time].value

    if sheet[row][aud].value:

        comp = int(str(sheet[row][aud].value)[0])

        if comp not in comps:
            comps[comp] = {}

        if d not in comps[comp]:
            comps[comp][d] = {}

        if lesson_time not in comps[comp][d]:
            comps[comp][d][lesson_time] = [gr]

        else:

            if gr not in comps[comp][d][lesson_time]:

                comps[comp][d][lesson_time].append(gr)

    if gr not in groups:
        groups[gr] = [{}, {}]

    if d not in groups[gr][wk-1]:
        groups[gr][wk-1][d] = [] 


    groups[gr][wk-1][d].append({

        "time": sheet[row][time].value,
        "subject": sheet[row][subject].value,
        "teacher": sheet[row][teacher].value,
        "place": sheet[row][aud].value

    })













# week - четность недели: 0 - нечётная, 1 - чётная

def date_schedule(group, date):

    nday = get_date(date).weekday()
    day = days[nday]

    return day_schedule(group, day, get_week_parity(date))


def day_schedule(group, day, week=get_week_parity(pendulum.now())):

    if not day in groups[str(group)][week]:

        return 0

    else:   

        schedule_for_day = groups[str(group)][week][day]

        return schedule_for_day

        # for less in schedule_for_day:

        #     print("    ", less['time'], less['subject'], less['place'], less['teacher'])


def week_schedule(group, week=get_week_parity(pendulum.now())):

    schedule_for_week = groups[str(group)][week]

    return schedule_for_week


def get_nums_groups(comp, dayweek, time):

    return len(comps[comp][dayweek][time])

def get_nums_pars(comp, dayweek):

    return len(comps[comp][dayweek])



